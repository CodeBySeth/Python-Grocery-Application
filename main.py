from openpyxl import Workbook, load_workbook
import os.path
import re

#Variables
answer        = ''
item          = ''
quantity      = 0
index         = 1
is_match      = False
file_path     = 'Grocery List.xlsx'
difference    = 1000
inventory_col = 'A'
quantity_col  = 'B'
recipe_col    = 'D'
reorder_col   = 'G'

#Find the empty cell in the column
def find_empty_cell(col_letter):
    index = 1
    for cell in ws[f"{col_letter}"]:
        if cell.value == None:
            break
        else:
            index += 1

    if index > ws.max_row:
        print('This is the index: ', index)
        return ws.max_row + 1
    else:
        print('This is the index: ', index)
        return index



if os.path.exists(file_path):
    wb = load_workbook('Grocery List.xlsx')

    ws = wb.active

    print('Would you like to add some items to your inventory?')

    while answer.capitalize() != 'N':
        answer = input('Enter Y for yes and N to continue: ')

        if answer.capitalize() == 'Y':
            print('Type in your items from your inventory and the quantity of items')
            print('If you are done adding items, type "stop" when in the item input.')

            while item != 'stop':
                item = input('Add an item: ')

                if item.lower() == 'stop' or item == '':
                    break

                else:
                    quantity = input('Quantity: ')
                    empty_cell = find_empty_cell(inventory_col)
                    ws[f"{inventory_col}{empty_cell}"] = item
                    ws[f"{quantity_col}{empty_cell}"] = quantity

                    #ws[f"{inventory_col}{ws.max_row+1}"] = item
                    #ws[f"{quantity_col}{ws.max_row}"] = quantity


    recipe = ''
    while recipe != "stop":
        recipe = input('Enter recipe: ')
        if recipe == 'stop':
            break
        matches      = re.findall(r'(?<!\d)(\d+/\d+|\d+)(?!\d)', recipe)
        quantity     = [1 if '/' in match else int(match) for match in matches]
        matches      = re.findall(r'\b[a-zA-Z]+\b', recipe)
        common_words = set(['and', 'or', 'of', 'to', 'for', 'with', 'optional', 'use', 'peeled', 'deveined', 'pound', 'large', 'teaspoon', 'dried', 'teaspoons', 'serving', 'wedges', 'wedge', 'stop', 'leaves', 'fresh', 'chopped', 'juice', 'minced', 'crushed', 'cup', 'cups', 'divided', 'melted'])
        recipe       = [match for match in matches if match.lower() not in common_words]
        recipe_str   = ' '.join(recipe)

        if not quantity:
            quantity.append(1)

        if not recipe_str.isspace():
            for row in ws[f"{recipe_col}"]:
                if row.value is None:
                    ws[f"{recipe_col}{index}"] = recipe_str
                    ws[f"E{index}"] = ' '.join(map(str, quantity))
                    flag = True
                    break
                index += 1

            if flag == False:
                ws[f"{recipe_col}{ws.max_row+1}"] = recipe_str
                ws[f"E{ws.max_row}"] = ', '.join(map(str, quantity))

            flag = False


    inv_col_str       = ''
    recipe_col_str    = ''
    starting_index    = find_empty_cell(inventory_col)
    inv_col_index     = 2
    recipe_col_index  = 1
    reorder_col_index = 2

    #Loop through recipe item list
    for recipe_row in ws[f"{recipe_col}"]:
        #Assign recipe string
        recipe_col_str = recipe_row.value

        for inv_row in ws[f"{inventory_col}"]:
            #Assign inventory string
            inv_col_str = inv_row.value

            #Check for matching strings
            if inv_col_str is not None and recipe_col_str is not None:
                print('Recipe String: ', recipe_row.value)
                if inv_col_str in recipe_col_str:
                    is_match = True
                    break

        #If there are no matching strings, insert recipe item in inventory list
        if is_match == False and recipe_col_str != ws[f"{recipe_col}{ws.min_row}"]:
            ws[f"{inventory_col}{starting_index}"] = recipe_col_str
            ws[f"{quantity_col}{starting_index}"] = 0
            starting_index += 1
            print("There were none")

        is_match = False




    inv_col_index     = 2
    recipe_col_index  = 2
    reorder_col_index = 2


    print('Reorder Loop \n\n\n\n\n\n\n\n')
    #Loop processing inventory column
    for cell in ws[f"{inventory_col}"]:
        inv_col_str = cell.value
        print('first loop: ', inv_col_index)
        recipe_col_index = 2

        #Loop processing recipe column
        for row in ws[f"{recipe_col}"]:
            recipe_col_str = row.value
            print('Recipe Col Index: ', recipe_col_index)
            #Check for first empty cell
            if recipe_col_str != None and inv_col_str != None:
                print('recipe: ' + recipe_col_str)
                print('inventory: ' + inv_col_str)
                #If inventory is in recipe then add to reorder
                if inv_col_str in recipe_col_str:

                    '''
                    print('inv index: ', inv_col_index - 1)
                    print('recipe index: ', recipe_col_index - 1)
                    print(ws[f"{quantity_col}{inv_col_index - 1}"].value)
                    print(ws[f"{'E'}{recipe_col_index - 1}"].value)'''


                    if type(ws[f"{quantity_col}{inv_col_index - 1}"].value) is str or type(ws[f"{'E'}{recipe_col_index - 1}"].value) == str:
                        difference = int(ws[f"{quantity_col}{inv_col_index - 1}"].value) - int(ws[f"{'E'}{recipe_col_index - 1}"].value)

                    if difference <= 0:
                        ws[f"{'H'}{reorder_col_index}"] = difference * -1
                        ws[f"{reorder_col}{reorder_col_index}"] = inv_col_str
                        print('added to reorder')
                recipe_col_index += 1

        inv_col_index += 1
    print('after loop')


    wb.save('Grocery List.xlsx')
    wb.close()
else:
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Inventory'
    ws['D1'] = 'Recipes'
    ws['G1'] = 'What To Order'

    wb.save('Grocery List.xlsx')